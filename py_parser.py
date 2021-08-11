# -*- coding: utf-8 -*-
# vim:fileencoding=utf-8
# encoding=utf8
import requests,ssl
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import openpyxl
import pandas as pd
import pymysql
import datetime
from PIL import Image
from bs4 import BeautifulSoup
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
#ssl._create_default_https_context = ssl._create_unverified_context
f=open(r'pythonscript/test1.xlsx',"wb")
ufr = requests.get("https://www.santech.ru/files/priceList/Каталог+товаров-11-sheets.xlsx",verify=False)
f.write(ufr.content)
f.close()

wb = openpyxl.load_workbook(filename = 'pythonscript/test1.xlsx')
url="https://www.santech.ru"
tb=True
id_category=1000
id_podcategory=4000
id_item=1
id_image=100
id_brand=1
id_attr=1
brands=[]
attributes=[]
mas_id_attr=[]
mas_id_brands=[]
def mysql_insert(args):
    args_str=""
    for arg in range(len(args)):
        if(arg==len(args)-1):
            args_str+="`"+args[arg]+"`"
        else:
            args_str+="`"+args[arg]+"`,"

    return args_str
def set_url_img(images,id_item,id_image):
    if(id_image=='0'):
        s="loadProduct/img"+str(id_item)+".jpg"
        image=''
        if(images!=''):
            p = requests.get(images,verify=False)
            out = open(r'beta.delight.su/public_html/image/loadProduct/img'+str(id_item)+'.jpg', "wb")
            out.write(p.content)
            out.close()
            try:
                im =Image.open(r'beta.delight.su/public_html/image/loadProduct/img'+str(id_item)+'.jpg')
                if(im.size!=0):
                    image = s
                else:
                    image="loadProduct/coming-soon.png"
            except:
                image="loadProduct/coming-soon.png"
        else:
            image="loadProduct/coming-soon.png"
    elif(id_image=="brand"):
        s="loadProduct/img"+str(id_item)+"_"+str(id_image)+".png"
        image=''
        if(images!=''):
            p = requests.get(images,verify=False)
            out = open(r'beta.delight.su/public_html/image/loadProduct/img'+str(id_item)+'_'+str(id_image)+'.png', "wb")
            out.write(p.content)
            out.close()
            try:
                im =Image.open(r'beta.delight.su/public_html/image/loadProduct/img'+str(id_item)+'_'+str(id_image)+'.png')
                if(im.size!=0):
                    image = s
                else:
                    image=''
            except:
                image=''
        else:
            image=''
    else:
        s="loadProduct/img"+str(id_item)+"_"+str(id_image)+".jpg"
        image=''
        if(images!=''):
            p = requests.get(images,verify=False)
            out = open(r'beta.delight.su/public_html/image/loadProduct/img'+str(id_item)+'_'+str(id_image)+'.jpg', "wb")
            out.write(p.content)
            out.close()
            try:
                im =Image.open(r'beta.delight.su/public_html/image/loadProduct/img'+str(id_item)+'_'+str(id_image)+'.jpg')
                if(im.size!=0):
                    image = s
                else:
                    image=''
            except:
                image=''
        else:
            image=''
    return image
def sear_cat(name):
    global wb,tb,id_category,id_podcategory,id_item,id_image,id_brand,pd,brands,mas_id_brands,mas_id_attr,id_attr
    connection = pymysql.connect(host='localhost',
                             user='h96568b5_ocar2',
                             password='E3J$4cgn%',
                             db='h96568b5_ocar2',
                             charset='utf8mb4'
                             )
    tables=["oc_category","oc_category_description","oc_category_to_store","oc_category_path","oc_category_to_layout","oc_product","oc_product_description",
            "oc_product_to_category","oc_product_image","oc_product_to_store","oc_manufacturer","oc_manufacturer_to_store",'oc_product_attribute','oc_attribute','oc_attribute_description']
    if(tb):
        tb=False
        for table in tables:
            cur = connection.cursor()
            cur.execute("TRUNCATE TABLE `"+table+"`")
    cur = connection.cursor()
    sheet=wb[name]
    for row in range(5,sheet.max_row-1):
        if(sheet.cell(row=row, column=1).font.size==16):
            v=sheet.cell(row=row, column=1).value
            cat_name=v
            if(v.count("-")!=0):
                cat_name=v[v.index("-")+2:]
            cur.execute("INSERT INTO `oc_category`(`category_id`, `image`,`parent_id`,`top`,`column`,`sort_order`,`status`,`date_added`,`date_modified`) VALUES (%s,NULL,%s,1,1,1,1,%s,%s)",
                      [id_category,0,str(datetime.datetime.now()),str(datetime.datetime.now())])
            cur.execute("INSERT INTO `oc_category_description`(`category_id`,`language_id`,`name`,`description`,`meta_title`,`meta_description`,`meta_keyword`) VALUES (%s,1,%s,'',%s,'','')",
                      [id_category,cat_name,cat_name])
            cur.execute("INSERT INTO `oc_category_to_store`(`category_id`,`store_id`) VALUES (%s,0)",
                      [id_category])
            cur.execute("INSERT INTO `oc_category_path`(`category_id`,`path_id`,`level`) VALUES (%s,%s,0)",
                      [id_category,id_category])
            cur.execute("INSERT INTO `oc_category_to_layout`(`category_id`,`store_id`,`layout_id`) VALUES (%s,0,0)",
                      [id_category])
            id_category+=1
        elif(sheet.cell(row=row, column=1).font.size==14):
            v=sheet.cell(row=row, column=1).value
            cat_name=v
            if(v.count("-")!=0):
                cat_name=v[v.index("-")+2:]
            cur.execute("INSERT INTO `oc_category`(`category_id`, `image`,`parent_id`,`top`,`column`,`sort_order`,`status`,`date_added`,`date_modified`) VALUES (%s,NULL,%s,1,1,1,1,%s,%s)",
                      [id_podcategory,id_category-1,str(datetime.datetime.now()),str(datetime.datetime.now())])
            cur.execute("INSERT INTO `oc_category_description`(`category_id`,`language_id`,`name`,`description`,`meta_title`,`meta_description`,`meta_keyword`) VALUES (%s,1,%s,'',%s,'','')",
                      [id_podcategory,cat_name,cat_name])
            cur.execute("INSERT INTO `oc_category_to_store`(`category_id`,`store_id`) VALUES (%s,0)",
                      [id_podcategory])
            cur.execute("INSERT INTO `oc_category_path`(`category_id`,`path_id`,`level`) VALUES (%s,%s,0)",
                      [id_podcategory,id_podcategory])
            cur.execute("INSERT INTO `oc_category_to_layout`(`category_id`,`store_id`,`layout_id`) VALUES (%s,0,0)",
                      [id_podcategory])
            id_podcategory+=1
        elif(sheet.cell(row=row, column=1).font.bold==False and sheet.cell(row=row, column=1).font.size==12):
            image="image/loadProduct/coming-soon.png"
            if(sheet.cell(row=row, column=7).hyperlink!=None):
                html_doc=sheet.cell(row=row, column=7).hyperlink.target
                tables=''
                try:
                    tables = pd.read_html(html_doc)
                except:
                    pass   
                name_attr=''
                attr=''
                for table in tables:
                    for rowtab in range(len(table.index)): 
                        name_attr=table.iloc[rowtab,0].encode('utf-8')
                        try:
                            attr_text=str(table.iloc[rowtab,1])
                        except:
                            attr_text=table.iloc[rowtab,1].encode('utf-8')

                        if(name_attr.count(name_attr.split(' ')[0])>1):
                            name_attr=name_attr[0:name_attr.index(name_attr.split(' ')[0],name_attr.index(name_attr.split(' ')[0])+1)-1]
                            if(attributes.count(name_attr)!=0):
                                attr=mas_id_attr[attributes.index(name_attr)]
                            else:
                                attributes.append(name_attr)
                                mas_id_attr.append(id_attr)
                                cur.execute("INSERT INTO `oc_attribute`("+mysql_insert(["attribute_id","attribute_group_id","sort_order"])+") VALUES (%s,%s,1)",
                                    [id_attr,3])#т.к 3-наименование
                                cur.execute("INSERT INTO `oc_attribute_description`("+mysql_insert(["attribute_id","language_id","name"])+") VALUES (%s,1,%s)",
                                    [id_attr,name_attr])#описание атрибута
                                attr=id_attr
                                id_attr+=1
                        else:
                            if(attributes.count(name_attr)!=0):
                                attr=mas_id_attr[attributes.index(name_attr)]
                            else:
                                attributes.append(name_attr)
                                mas_id_attr.append(id_attr)
                                cur.execute("INSERT INTO `oc_attribute`("+mysql_insert(["attribute_id","attribute_group_id","sort_order"])+") VALUES (%s,%s,1)",
                                    [id_attr,3])#т.к 3-наименование
                                cur.execute("INSERT INTO `oc_attribute_description`("+mysql_insert(["attribute_id","language_id","name"])+") VALUES (%s,1,%s)",
                                    [id_attr,name_attr])#описание атрибута
                                attr=id_attr
                                id_attr+=1
                        try:
                            cur.execute("INSERT INTO `oc_product_attribute`("+mysql_insert(["product_id","attribute_id","language_id","text"])+") VALUES (%s,%s,1,%s)",
                            [id_item,attr,attr_text])
                        except:
                            cur.execute("INSERT INTO `oc_product_attribute`("+mysql_insert(["product_id","attribute_id","language_id","text"])+") VALUES (%s,%s,0,%s)",
                            [id_item,attr,attr_text])
                   
                soup=''   
                try:
                    response = requests.get(html_doc,verify=False)
                    soup = BeautifulSoup (response.text, 'html.parser')
                except:
                    pass
                images=''
                im1=0
                
                for link in soup.find_all('a', class_="gallery__item-gal"):
                    if im1==0:
                        images=url+link.get('href')
                        im1=1
                        image=set_url_img(images.encode('utf-8'),id_item,'0')
                    else:
                        images=set_url_img(url+link.get('href').encode('utf-8'),id_item,id_image.encode('utf-8'))
                        if(images!=''):
                            cur.execute("INSERT INTO `oc_product_image`("+mysql_insert(["product_image_id","product_id","image","sort_order"])+") VALUES (%s,%s,%s,0)",
                                [id_image,id_item,images])
                            id_image+=1
                    
            else:
                image="loadProduct/coming-soon.png"
                
            cur = connection.cursor()
            name=sheet.cell(row=row, column=4).value.encode('utf-8')
            price=sheet.cell(row=row, column=9).value
            code=''
            if(sheet.cell(row=row, column=1).value!=None):
                code=sheet.cell(row=row, column=1).value
            article=''
            if(sheet.cell(row=row, column=3).value!=None):
                article=sheet.cell(row=row, column=3).value
            manufacturer=''
            if(sheet.cell(row=row, column=6).value!=None):
                if(brands.count(sheet.cell(row=row, column=6).value)!=0):
                    manufacturer=mas_id_brands[brands.index(sheet.cell(row=row, column=6).value)]
                else:
                    brands.append(sheet.cell(row=row, column=6).value)
                    mas_id_brands.append(id_brand)
                    brand_icon=''
                    for link_icon in soup.find_all('img', class_="variants__brand-img"):
                        brand_icon=url+link_icon.get('src')
                    if(brand_icon!=''):
                        
                        cur.execute("INSERT INTO `oc_manufacturer`("+mysql_insert(["manufacturer_id","name","image","sort_order"])+") VALUES (%s,%s,%s,0)",
                            [id_brand,sheet.cell(row=row, column=6).value,set_url_img(brand_icon.encode('utf-8'),id_brand,'brand')])
                    else:
                        cur.execute("INSERT INTO `oc_manufacturer`("+mysql_insert(["manufacturer_id","name","image","sort_order"])+") VALUES (%s,%s,%s,0)",
                            [id_brand,sheet.cell(row=row, column=6).value,""])
                    cur.execute("INSERT INTO `oc_manufacturer_to_store`("+mysql_insert(["manufacturer_id","store_id"])+") VALUES (%s,%s)",
                            [id_brand,0])
                    manufacturer=id_brand
                    id_brand+=1
            cur.execute("INSERT INTO `oc_product`("+mysql_insert(["product_id","model","sku","upc","ean","jan","isbn","mpn","location","quantity","stock_status_id","image","manufacturer_id","shipping","price","points","tax_class_id","date_available","weight","weight_class_id","length","width","height","length_class_id","subtract","minimum","sort_order","status","viewed","date_added","date_modified"])+") VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                            [id_item,code,article,"","","","","","",50,6,image,int(manufacturer),1,price,0,0,'',0.0000,0,0.0000,0.0000,0.0000,0,1,1,0,1,0,str(datetime.datetime.now()),str(datetime.datetime.now())])
            cur.execute("INSERT INTO `oc_product_description`("+mysql_insert(["product_id","language_id","name","description","tag","meta_title","meta_description","meta_keyword"])+") VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                            [id_item,1,name,"","",name,"",""])
            cur.execute("INSERT INTO `oc_product_to_category`("+mysql_insert(["product_id","category_id"])+") VALUES (%s,%s)",
                            [id_item,id_category-1])
            cur.execute("INSERT INTO `oc_product_to_category`("+mysql_insert(["product_id","category_id"])+") VALUES (%s,%s)",
                            [id_item,id_podcategory-1])
            cur.execute("INSERT INTO `oc_product_to_store`("+mysql_insert(["product_id","store_id"])+") VALUES (%s,%s)",
                            [id_item,0])
            id_item+=1
for i in range(1,len(wb.sheetnames)):
    if(u"Распродажа"==wb.sheetnames[i]):
        continue
    sear_cat(wb.sheetnames[i])
    
